"""
GST Billing Software
Features: Logo in PDF header, Excel ledger, Google Sheets live sync
Run: python billing_app.py
Requirements: pip install openpyxl pillow reportlab gspread google-auth
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
import os, json, copy, threading
from PIL import Image, ImageTk

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                 Paragraph, Spacer)
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
from reportlab.lib.utils import ImageReader   # ← correct way to embed images in PDF

# ── File paths (always relative to this script's folder) ──────────────────────
_DIR        = os.path.dirname(os.path.abspath(__file__))
CONFIG_FILE = os.path.join(_DIR, "company_config.json")
BILLS_FILE  = os.path.join(_DIR, "all_bills.xlsx")
LOGO_FILE   = os.path.join(_DIR, "logo.jpg")
CREDS_FILE  = os.path.join(_DIR, "google_credentials.json")   # Service account JSON from Google

# ── Brand colours ──────────────────────────────────────────────────────────────
GREEN_DARK  = "#1a6b3c"
GREEN_MID   = "#2e8b57"
GREEN_LIGHT = "#d4edda"
ORANGE      = "#e87722"
BG          = "#f4fbf6"
ROW_ALT     = "#e8f5e9"

# ── Default config ────────────────────────────────────
DEFAULT_CONFIG = {
    "company_name":  "Your Company Name",
    "legal_name":    "Legal Owner Name",
    "trade_name":    "Trade Name",
    "address":       "Your Address, City, State - PIN",
    "phone":         "",
    "email":         "",
    "gstin":         "",
    "udyam":         "",
    "bank_name":     "Your Bank Name",
    "account_name":  "Your Company Name",
    "account_no":    "",
    "account_type":  "Current A/C",
    "ifsc":          "",
    "branch":        "Branch Name (Branch Code: XXXXXX)",
    "logo_path":     os.path.join(os.path.dirname(os.path.abspath(__file__)), "logo.jpg"),  # Place your logo.jpg here
    "cgst_rate":     "2.5",
    "sgst_rate":     "2.5",
    # Google Sheets settings
    "gsheet_enabled": False,
    "gsheet_id":      "",        # Paste your Google Sheet ID here
    "gsheet_tab":     "Bills",   # Sheet tab name
}

def load_config():
    if os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE) as f:
            cfg = json.load(f)
        for k, v in DEFAULT_CONFIG.items():
            cfg.setdefault(k, v)
        return cfg
    return copy.deepcopy(DEFAULT_CONFIG)

def save_config(cfg):
    with open(CONFIG_FILE, "w") as f:
        json.dump(cfg, f, indent=2)

def next_bill_no():
    if os.path.exists(BILLS_FILE):
        try:
            wb = openpyxl.load_workbook(BILLS_FILE, read_only=True)
            ws = wb["Bills"]
            count = sum(1 for row in ws.iter_rows(min_row=2, values_only=True) if row[0])
            wb.close()
            return 1001 + count
        except Exception:
            pass
    return 1001

# ── Number to words (Indian) ───────────────────────────────────────────────────
def num_to_words(n):
    ones = ['','One','Two','Three','Four','Five','Six','Seven','Eight','Nine',
            'Ten','Eleven','Twelve','Thirteen','Fourteen','Fifteen','Sixteen',
            'Seventeen','Eighteen','Nineteen']
    tens = ['','','Twenty','Thirty','Forty','Fifty','Sixty','Seventy','Eighty','Ninety']
    def h(x):
        if x == 0: return ''
        if x < 20: return ones[x]
        if x < 100: return tens[x//10]+(' '+ones[x%10] if x%10 else '')
        return ones[x//100]+' Hundred'+(' '+h(x%100) if x%100 else '')
    n = int(n)
    if n == 0: return 'Zero'
    parts = []
    for div, lbl in [(10000000,'Crore'),(100000,'Lakh'),(1000,'Thousand'),(1,'')]:
        q = n // div; n %= div
        if q: parts.append(h(q)+(' '+lbl if lbl else ''))
    return ' '.join(parts).strip() + ' Rupees Only'

# ── Excel ledger ───────────────────────────────────────────────────────────────
HEADERS = ["Bill No","Date","Customer Name","Address","Phone","GSTIN",
           "Place of Supply","Items Description","HSN/SAC","Qty","Rate",
           "Taxable Value","CGST %","CGST Amt","SGST %","SGST Amt","Total"]

def save_bill_to_excel(d):
    """Save bill row to local Excel file."""
    if os.path.exists(BILLS_FILE):
        wb = openpyxl.load_workbook(BILLS_FILE)
        ws = wb["Bills"]
    else:
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Bills"
        # Header row
        for c, h in enumerate(HEADERS, 1):
            cell = ws.cell(1, c, h)
            cell.font      = Font(bold=True, color="FFFFFF", name="Arial")
            cell.fill      = PatternFill("solid", start_color="1a6b3c")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 28
        # Summary sheet
        ws2 = wb.create_sheet("Summary")
        for r,(lbl,fml) in enumerate([
            ("Total Bills",        "=COUNTA(Bills!A2:A9999)"),
            ("Total Revenue (Rs.)","=SUM(Bills!Q2:Q9999)"),
            ("Total Tax (Rs.)",    "=SUM(Bills!N2:N9999)+SUM(Bills!P2:P9999)"),
        ], 1):
            ws2.cell(r,1,lbl).font = Font(bold=True, name="Arial")
            ws2.cell(r,2,fml).number_format = '#,##0.00'

    row_data = [
        d["bill_no"], d["date"], d["customer_name"], d["address"], d["phone"],
        d["gstin"], d["place_of_supply"], d["items_desc"], d["hsn"],
        d["qty"], d["rate"], d["taxable_value"],
        d["cgst_pct"], d["cgst_amt"], d["sgst_pct"], d["sgst_amt"], d["total"]
    ]
    ws.append(row_data)

    # Format currency columns
    for col in [11,12,14,16,17]:
        ws.cell(ws.max_row, col).number_format = '#,##0.00'
    # Alternating row colour
    if ws.max_row % 2 == 0:
        fill = PatternFill("solid", start_color="e8f5e9")
        for c in range(1, len(HEADERS)+1):
            ws.cell(ws.max_row, c).fill = fill

    col_w = [8,12,22,32,14,20,18,40,10,6,14,14,8,12,8,12,14]
    for i,w in enumerate(col_w,1):
        ws.column_dimensions[get_column_letter(i)].width = w
    wb.save(BILLS_FILE)
    return row_data

# ── Google Sheets sync ─────────────────────────────────────────────────────────
def sync_to_gsheet(cfg, row_data, status_callback=None):
    """Push one bill row to Google Sheets. Runs in background thread."""
    if not cfg.get("gsheet_enabled") or not cfg.get("gsheet_id"):
        return
    if not os.path.exists(CREDS_FILE):
        if status_callback:
            status_callback("⚠ google_credentials.json not found – skipping sync")
        return
    try:
        import gspread
        from google.oauth2.service_account import Credentials
        scopes = ["https://spreadsheets.google.com/feeds",
                  "https://www.googleapis.com/auth/drive"]
        creds = Credentials.from_service_account_file(CREDS_FILE, scopes=scopes)
        gc    = gspread.authorize(creds)
        sh    = gc.open_by_key(cfg["gsheet_id"])
        tab_name = cfg.get("gsheet_tab", "Bills")

        # Get or create tab
        try:
            ws = sh.worksheet(tab_name)
        except gspread.WorksheetNotFound:
            ws = sh.add_worksheet(title=tab_name, rows=2000, cols=20)
            ws.append_row(HEADERS)
            # Format header (bold + green background via batch update)

        ws.append_row([str(v) for v in row_data])
        if status_callback:
            status_callback("✅ Google Sheets updated!")
    except Exception as e:
        if status_callback:
            status_callback(f"⚠ Google Sheets error: {e}")

# ══════════════════════════════════════════════════════════════════════════════
#  PDF GENERATOR  (logo via ImageReader – works 100%)
# ══════════════════════════════════════════════════════════════════════════════
def pstyle(name, size, bold=False, align=TA_LEFT, color=colors.black):
    return ParagraphStyle(name, fontSize=size, leading=size+2.5,
                          fontName='Helvetica-Bold' if bold else 'Helvetica',
                          alignment=align, textColor=color)

def generate_pdf(d, cfg, filepath):
    doc = SimpleDocTemplate(filepath, pagesize=A4,
                            topMargin=8*mm, bottomMargin=8*mm,
                            leftMargin=10*mm, rightMargin=10*mm)
    W = 190*mm
    elems = []

    n8  = pstyle('n8', 8)
    b8  = pstyle('b8', 8, True)
    b9  = pstyle('b9', 9, True)
    r8  = pstyle('r8', 8, align=TA_RIGHT)
    c8  = pstyle('c8', 8, align=TA_CENTER)
    qt  = pstyle('qt',14, True, align=TA_CENTER, color=colors.HexColor('#1a6b3c'))

    # ── LOGO using ImageReader (correct ReportLab method) ──────────────────────
    from reportlab.platypus import Image as RLImg
    logo_path = cfg.get("logo_path", "")
    logo_cell = ""
    if logo_path and os.path.exists(logo_path):
        try:
            # Resize to max 28mm wide keeping aspect ratio
            pil_img = Image.open(logo_path).convert("RGB")
            orig_w, orig_h = pil_img.size
            target_w_px = int(28 * 3.779527)   # 28mm in pixels at 96dpi
            scale = target_w_px / orig_w
            new_h_mm = (orig_h * scale) / 3.779527
            logo_cell = RLImg(logo_path,
                              width=28*mm,
                              height=min(new_h_mm, 26)*mm)
        except Exception as e:
            logo_cell = ""

    # ── Header row ─────────────────────────────────────────────────────────────
    co_text = Paragraph(
        f'<font size="16" color="#1a6b3c"><b>{cfg["company_name"]}</b></font><br/>'
        f'<font size="8">Prop: <b>{cfg["legal_name"]}</b>  |  Trade: {cfg["trade_name"]}<br/>'
        f'{cfg["address"]}'
        f'{"<br/>Ph: "+cfg["phone"] if cfg["phone"] else ""}'
        f'{"  |  "+cfg["email"] if cfg["email"] else ""}'
        f'</font>', n8)

    ri_text = Paragraph(
        f'<font size="8">'
        f'<b>GSTIN:</b> {cfg["gstin"]}<br/>'
        f'{"<b>UDYAM:</b> "+cfg["udyam"]+"<br/>" if cfg["udyam"] else ""}'
        f'<b>Bank:</b> {cfg["bank_name"]}<br/>'
        f'<b>A/C No:</b> {cfg["account_no"]} ({cfg["account_type"]})<br/>'
        f'<b>IFSC:</b> {cfg["ifsc"]}<br/>'
        f'<b>Branch:</b> {cfg["branch"]}'
        f'</font>', n8)

    if logo_cell:
        hdr_data = [[logo_cell, co_text, ri_text]]
        hdr_cols = [30*mm, 95*mm, 65*mm]
    else:
        hdr_data = [[co_text, ri_text]]
        hdr_cols = [125*mm, 65*mm]

    hdr = Table(hdr_data, colWidths=hdr_cols)
    hdr.setStyle(TableStyle([
        ('VALIGN',       (0,0), (-1,-1), 'MIDDLE'),
        ('LINEBELOW',    (0,0), (-1,-1), 2, colors.HexColor('#1a6b3c')),
        ('BOTTOMPADDING',(0,0), (-1,-1), 5),
        ('TOPPADDING',   (0,0), (-1,-1), 3),
        ('LEFTPADDING',  (0,0), (-1,-1), 3),
    ]))
    elems.append(hdr)
    elems.append(Spacer(1, 3*mm))

    # ── Quotation title bar ────────────────────────────────────────────────────
    tt = Table([[
        Paragraph(f'<b>GSTIN: {cfg["gstin"]}</b>', n8),
        Paragraph('<b>QUOTATION</b>', qt),
        Paragraph('ORIGINAL FOR RECIPIENT', r8),
    ]], colWidths=[65*mm, 65*mm, 60*mm])
    tt.setStyle(TableStyle([
        ('BACKGROUND',    (0,0), (-1,-1), colors.HexColor('#d4edda')),
        ('BOX',           (0,0), (-1,-1), 1,   colors.HexColor('#1a6b3c')),
        ('TOPPADDING',    (0,0), (-1,-1), 5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ('VALIGN',        (0,0), (-1,-1), 'MIDDLE'),
    ]))
    elems.append(tt)
    elems.append(Spacer(1, 2*mm))

    # ── Customer block ─────────────────────────────────────────────────────────
    cust = [
        ['Customer Detail', '', 'Quotation No.', str(d["bill_no"]), 'Date', d["date"]],
        ['Name',    d["customer_name"], '', '', '', ''],
        ['Address', d["address"],       '', '', '', ''],
        ['Phone',   d["phone"],          '', '', '', ''],
        ['GSTIN',   d["gstin"] or '-',  '', '', '', ''],
        ['Place of Supply', d["place_of_supply"], '', '', '', ''],
    ]
    ct = Table(cust, colWidths=[28*mm, 67*mm, 28*mm, 22*mm, 22*mm, 23*mm])
    ct.setStyle(TableStyle([
        ('SPAN',         (0,0), (1,0)),
        ('BACKGROUND',   (0,0), (-1,0),  colors.HexColor('#d4edda')),
        ('FONTNAME',     (0,0), (-1,0),  'Helvetica-Bold'),
        ('FONTSIZE',     (0,0), (-1,-1), 8),
        ('BOX',          (0,0), (-1,-1), 0.5, colors.black),
        ('INNERGRID',    (0,0), (-1,-1), 0.25, colors.HexColor('#aaaaaa')),
        ('BACKGROUND',   (0,1), (0,-1),  colors.HexColor('#f2f9f4')),
        ('FONTNAME',     (0,1), (0,-1),  'Helvetica-Bold'),
        ('FONTNAME',     (2,0), (2,-1),  'Helvetica-Bold'),
        ('FONTNAME',     (4,0), (4,-1),  'Helvetica-Bold'),
        ('VALIGN',       (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING',   (0,0), (-1,-1), 2),
        ('BOTTOMPADDING',(0,0), (-1,-1), 2),
    ]))
    elems.append(ct)
    elems.append(Spacer(1, 2*mm))

    # ── Items table ────────────────────────────────────────────────────────────
    desc_p = Paragraph(d["items_desc"].replace('\n', '<br/>'), n8)
    it_rows = [
        ['Sr.', 'Name of Product / Service', 'HSN/SAC', 'Qty', 'Rate (Rs.)', 'Taxable Value (Rs.)'],
        ['1', desc_p, d["hsn"], str(d["qty"]),
         f'{d["rate"]:,.2f}', f'{d["taxable_value"]:,.2f}'],
        ['','','','', f'CGST ({d["cgst_pct"]}%)', f'{d["cgst_amt"]:,.2f}'],
        ['','','','', f'SGST ({d["sgst_pct"]}%)', f'{d["sgst_amt"]:,.2f}'],
        ['','','','', 'TOTAL', f'{d["total"]:,.2f}'],
    ]
    it = Table(it_rows, colWidths=[10*mm, 90*mm, 20*mm, 12*mm, 28*mm, 30*mm])
    it.setStyle(TableStyle([
        ('BACKGROUND',   (0,0),  (-1,0),  colors.HexColor('#1a6b3c')),
        ('TEXTCOLOR',    (0,0),  (-1,0),  colors.white),
        ('FONTNAME',     (0,0),  (-1,0),  'Helvetica-Bold'),
        ('FONTSIZE',     (0,0),  (-1,-1), 8),
        ('BOX',          (0,0),  (-1,-1), 0.5, colors.black),
        ('INNERGRID',    (0,0),  (-1,-1), 0.25, colors.HexColor('#aaaaaa')),
        ('ALIGN',        (2,0),  (-1,-1), 'RIGHT'),
        ('VALIGN',       (0,0),  (-1,-1), 'TOP'),
        ('ROWBACKGROUNDS',(0,1), (-1,-2), [colors.white, colors.HexColor('#e8f5e9')]),
        ('BACKGROUND',   (0,-1), (-1,-1), colors.HexColor('#1a6b3c')),
        ('TEXTCOLOR',    (0,-1), (-1,-1), colors.white),
        ('FONTNAME',     (0,-1), (-1,-1), 'Helvetica-Bold'),
        ('TOPPADDING',   (0,0),  (-1,-1), 3),
        ('BOTTOMPADDING',(0,0),  (-1,-1), 3),
    ]))
    elems.append(it)
    elems.append(Spacer(1, 2*mm))

    # ── Total in words ─────────────────────────────────────────────────────────
    wt = Table([[Paragraph(f'<b>Total in words:</b>  {num_to_words(d["total"])}', n8)]],
               colWidths=[W])
    wt.setStyle(TableStyle([
        ('BOX',          (0,0), (-1,-1), 0.5, colors.black),
        ('BACKGROUND',   (0,0), (-1,-1), colors.HexColor('#f9fdf9')),
        ('TOPPADDING',   (0,0), (-1,-1), 3),
        ('BOTTOMPADDING',(0,0), (-1,-1), 3),
    ]))
    elems.append(wt)
    elems.append(Spacer(1, 2*mm))

    # ── Tax summary table ──────────────────────────────────────────────────────
    tx_rows = [
        ['HSN/SAC','Taxable Value','CGST','','SGST','','Total Tax'],
        ['','','%','Amount','%','Amount',''],
        [d["hsn"], f'Rs.{d["taxable_value"]:,.2f}',
         str(d["cgst_pct"]), f'Rs.{d["cgst_amt"]:,.2f}',
         str(d["sgst_pct"]), f'Rs.{d["sgst_amt"]:,.2f}',
         f'Rs.{d["total_tax"]:,.2f}'],
        ['Total', f'Rs.{d["taxable_value"]:,.2f}', '',
         f'Rs.{d["cgst_amt"]:,.2f}', '',
         f'Rs.{d["sgst_amt"]:,.2f}', f'Rs.{d["total_tax"]:,.2f}'],
    ]
    xt = Table(tx_rows, colWidths=[25*mm,35*mm,15*mm,28*mm,15*mm,28*mm,24*mm])
    xt.setStyle(TableStyle([
        ('SPAN',         (2,0),  (3,0)),
        ('SPAN',         (4,0),  (5,0)),
        ('BACKGROUND',   (0,0),  (-1,1),  colors.HexColor('#d4edda')),
        ('FONTNAME',     (0,0),  (-1,1),  'Helvetica-Bold'),
        ('FONTSIZE',     (0,0),  (-1,-1), 8),
        ('BOX',          (0,0),  (-1,-1), 0.5, colors.black),
        ('INNERGRID',    (0,0),  (-1,-1), 0.25, colors.HexColor('#aaaaaa')),
        ('ALIGN',        (1,0),  (-1,-1), 'RIGHT'),
        ('BACKGROUND',   (0,-1), (-1,-1), colors.HexColor('#f2f9f4')),
        ('FONTNAME',     (0,-1), (-1,-1), 'Helvetica-Bold'),
        ('TOPPADDING',   (0,0),  (-1,-1), 2),
        ('BOTTOMPADDING',(0,0),  (-1,-1), 2),
    ]))
    elems.append(xt)

    twtx = Table([[Paragraph(f'<b>Total Tax in words:</b>  {num_to_words(d["total_tax"])}', n8)]],
                 colWidths=[W])
    twtx.setStyle(TableStyle([
        ('BOX',          (0,0), (-1,-1), 0.5, colors.black),
        ('TOPPADDING',   (0,0), (-1,-1), 3),
        ('BOTTOMPADDING',(0,0), (-1,-1), 3),
    ]))
    elems.append(twtx)
    elems.append(Spacer(1, 2*mm))

    # ── Bank + Signature ───────────────────────────────────────────────────────
    bk = Paragraph(
        f'<b>Bank Details</b><br/>'
        f'Bank Name    : {cfg["bank_name"]}<br/>'
        f'Account Name : {cfg["account_name"]}<br/>'
        f'Account No.  : {cfg["account_no"]} ({cfg["account_type"]})<br/>'
        f'IFSC Code    : {cfg["ifsc"]}<br/>'
        f'Branch       : {cfg["branch"]}', n8)
    sg = Paragraph(
        'Certified that the particulars given above are true and correct.'
        '<br/><br/><br/>'
        f'<b>For {cfg["company_name"]}</b><br/>'
        f'<font size="7">(Prop: {cfg["legal_name"]})</font>', n8)

    bt = Table([[bk, sg]], colWidths=[110*mm, 80*mm])
    bt.setStyle(TableStyle([
        ('BOX',          (0,0), (-1,-1), 0.5, colors.black),
        ('INNERGRID',    (0,0), (-1,-1), 0.5, colors.black),
        ('TOPPADDING',   (0,0), (-1,-1), 4),
        ('BOTTOMPADDING',(0,0), (-1,-1), 4),
        ('VALIGN',       (0,0), (-1,-1), 'TOP'),
    ]))
    elems.append(bt)
    elems.append(Spacer(1, 2*mm))

    # ── Terms ──────────────────────────────────────────────────────────────────
    terms_html = d.get("terms","").replace('\n','<br/>')
    tct = Table([
        [Paragraph('<b>Terms and Conditions</b>', b9)],
        [Paragraph(terms_html, n8)],
    ], colWidths=[W])
    tct.setStyle(TableStyle([
        ('BOX',          (0,0), (-1,-1), 0.5, colors.black),
        ('BACKGROUND',   (0,0), (-1,0),  colors.HexColor('#d4edda')),
        ('TOPPADDING',   (0,0), (-1,-1), 3),
        ('BOTTOMPADDING',(0,0), (-1,-1), 3),
    ]))
    elems.append(tct)
    elems.append(Spacer(1, 3*mm))

    # ── Footer ─────────────────────────────────────────────────────────────────
    ft = Table([[Paragraph(
        f'<font size="7" color="#888888">'
        f'{cfg["company_name"]} | GSTIN: {cfg["gstin"]} | '
        f'Subject to Local Jurisdiction | Computer Generated Quotation'
        f'</font>', pstyle('cf',7,align=TA_CENTER))]], colWidths=[W])
    ft.setStyle(TableStyle([('LINEABOVE',(0,0),(-1,-1),0.5,colors.HexColor('#1a6b3c'))]))
    elems.append(ft)

    doc.build(elems)

# ══════════════════════════════════════════════════════════════════════════════
#  MAIN GUI APPLICATION
# ══════════════════════════════════════════════════════════════════════════════
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.cfg = load_config()
        self.title("GST Billing – " + self.cfg["company_name"])
        self.geometry("1040x860")
        self.configure(bg=BG)
        self.resizable(True, True)
        self._apply_style()
        self._build_header()

        nb = ttk.Notebook(self)
        nb.pack(fill="both", expand=True, padx=10, pady=5)

        self.tab_bill    = ttk.Frame(nb); nb.add(self.tab_bill,    text="  📄  New Bill  ")
        self.tab_company = ttk.Frame(nb); nb.add(self.tab_company, text="  🏢  Company Settings  ")
        self.tab_gsheet  = ttk.Frame(nb); nb.add(self.tab_gsheet,  text="  📡  Google Sheets  ")
        self.tab_ledger  = ttk.Frame(nb); nb.add(self.tab_ledger,  text="  📊  Bills Ledger  ")

        self._build_bill_tab()
        self._build_company_tab()
        self._build_gsheet_tab()
        self._build_ledger_tab()

    # ── ttk Style ──────────────────────────────────────────────────────────────
    def _apply_style(self):
        s = ttk.Style(self); s.theme_use("clam")
        s.configure("TNotebook",         background=BG)
        s.configure("TNotebook.Tab",     font=("Arial",10,"bold"), padding=[14,6],
                    background="#c8e6c9")
        s.map("TNotebook.Tab",           background=[("selected",GREEN_DARK)],
                                         foreground=[("selected","white")])
        s.configure("TFrame",            background=BG)
        s.configure("TLabelframe",       background=BG)
        s.configure("TLabelframe.Label", font=("Arial",10,"bold"),
                    foreground=GREEN_DARK, background=BG)
        s.configure("Treeview.Heading",  font=("Arial",9,"bold"),
                    background=GREEN_DARK, foreground="white")
        s.configure("Treeview",          font=("Arial",9), rowheight=22)

    # ── App Header (with logo thumbnail) ───────────────────────────────────────
    def _build_header(self):
        hf = tk.Frame(self, bg=GREEN_DARK, height=78)
        hf.pack(fill="x"); hf.pack_propagate(False)

        logo_path = self.cfg.get("logo_path","")
        if logo_path and os.path.exists(logo_path):
            try:
                img = Image.open(logo_path).convert("RGBA" if logo_path.endswith('.png') else "RGB")
                img = img.resize((60,48), Image.LANCZOS)
                self._logo_img = ImageTk.PhotoImage(img)
                tk.Label(hf, image=self._logo_img, bg=GREEN_DARK).pack(
                    side="left", padx=12, pady=6)
            except: pass

        lf = tk.Frame(hf, bg=GREEN_DARK); lf.pack(side="left", padx=4)
        tk.Label(lf, text=self.cfg["company_name"],
                 font=("Arial",21,"bold"), bg=GREEN_DARK, fg=ORANGE).pack(anchor="w")
        tk.Label(lf, text=f'Prop: {self.cfg["legal_name"]}  |  GSTIN: {self.cfg["gstin"]}',
                 font=("Arial",8), bg=GREEN_DARK, fg="#b7e4c7").pack(anchor="w")
        tk.Label(hf, text=self.cfg["address"],
                 font=("Arial",8), bg=GREEN_DARK, fg="#b7e4c7").pack(side="right", padx=20)

    # ── Helper: labelled entry ─────────────────────────────────────────────────
    def _lentry(self, parent, label, row, col=0, width=28, default="", bg=None):
        tk.Label(parent, text=label, font=("Arial",9,"bold"),
                 bg=bg or BG, fg=GREEN_DARK, anchor="w").grid(
                     row=row, column=col, sticky="w", padx=6, pady=3)
        e = tk.Entry(parent, width=width, font=("Arial",10),
                     relief="solid", bd=1, bg="white")
        e.grid(row=row, column=col+1, sticky="ew", padx=6, pady=3)
        if default: e.insert(0, default)
        return e

    # ══════════════════════════════════════════════════════════════════════════
    #  NEW BILL TAB
    # ══════════════════════════════════════════════════════════════════════════
    def _build_bill_tab(self):
        f = self.tab_bill
        canvas = tk.Canvas(f, bg=BG, highlightthickness=0)
        sb = ttk.Scrollbar(f, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y"); canvas.pack(side="left", fill="both", expand=True)
        inner = tk.Frame(canvas, bg=BG)
        win = canvas.create_window((0,0), window=inner, anchor="nw")
        inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>", lambda e: canvas.itemconfig(win, width=e.width))
        canvas.bind_all("<MouseWheel>",
                        lambda e: canvas.yview_scroll(int(-1*(e.delta/120)), "units"))

        # ── Customer Section ───────────────────────────────────────────────────
        cust = ttk.LabelFrame(inner, text="  Customer Details  ", padding=10)
        cust.pack(fill="x", padx=15, pady=8)
        for c in [1,3]: cust.columnconfigure(c, weight=1)

        self.e_cname  = self._lentry(cust, "Customer Name *", 0, 0, 30)
        self.e_phone  = self._lentry(cust, "Phone",           0, 2, 22)
        self.e_addr   = self._lentry(cust, "Address *",       1, 0, 30)
        self.e_cgstin = self._lentry(cust, "Customer GSTIN",  1, 2, 22)
        self.e_pos    = self._lentry(cust, "Place of Supply", 2, 0, 30,
                                     default="Your State (Code)")
        self.e_date   = self._lentry(cust, "Date",            2, 2, 22,
                                     default=datetime.today().strftime("%d-%m-%Y"))

        tk.Label(cust, text="Bill No.", font=("Arial",9,"bold"),
                 bg=BG, fg=GREEN_DARK).grid(row=3, column=0, sticky="w", padx=6, pady=3)
        self.e_billno = tk.Entry(cust, width=12, font=("Arial",13,"bold"),
                                  relief="solid", bd=1, bg="#e8f5e9", fg=GREEN_DARK)
        self.e_billno.insert(0, str(next_bill_no()))
        self.e_billno.grid(row=3, column=1, sticky="w", padx=6)

        # ── Product Section ────────────────────────────────────────────────────
        item = ttk.LabelFrame(inner, text="  Product / Service Details  ", padding=10)
        item.pack(fill="x", padx=15, pady=8)
        for c in [1,3]: item.columnconfigure(c, weight=1)

        tk.Label(item, text="Description *", font=("Arial",9,"bold"),
                 bg=BG, fg=GREEN_DARK).grid(row=0, column=0, sticky="nw", padx=6, pady=3)
        self.txt_desc = tk.Text(item, width=70, height=9, font=("Arial",9),
                                 relief="solid", bd=1, wrap="word", bg="white")
        self.txt_desc.grid(row=0, column=1, columnspan=3, sticky="ew", padx=6, pady=3)
        self.txt_desc.insert("1.0",
            "Enter product / service description here...")

        self.e_hsn  = self._lentry(item, "HSN / SAC",  1, 0, 16)
        self.e_qty  = self._lentry(item, "Qty",         1, 2, 10, default="1.00")
        self.e_rate = self._lentry(item, "Rate (Rs.)",  2, 0, 16)
        self.e_cgst = self._lentry(item, "CGST %",      2, 2, 10,
                                   default=self.cfg["cgst_rate"])
        self.e_sgst = self._lentry(item, "SGST %",      3, 0, 10,
                                   default=self.cfg["sgst_rate"])

        tk.Button(item, text="  Calculate  ", font=("Arial",10,"bold"),
                  bg=GREEN_DARK, fg="white", relief="flat", padx=10, pady=5,
                  cursor="hand2", command=self._calc).grid(row=3, column=2, padx=8, pady=6)

        # Results bar
        res = tk.Frame(item, bg="#e8f5e9", relief="groove", bd=1)
        res.grid(row=4, column=0, columnspan=4, sticky="ew", padx=6, pady=6)
        for col,(lbl,attr) in enumerate([
            ("Taxable Value","lbl_tax"), ("+ CGST","lbl_cgst"),
            ("+ SGST","lbl_sgst"),      ("= TOTAL","lbl_total")
        ]):
            tk.Label(res, text=lbl, font=("Arial",9,"bold"),
                     bg="#e8f5e9", fg=GREEN_DARK).grid(row=0, column=col*2, padx=12, pady=8)
            v = tk.Label(res, text="Rs. 0.00",
                         font=("Arial",13 if lbl=="= TOTAL" else 10, "bold"),
                         bg="#e8f5e9",
                         fg="#c00000" if lbl=="= TOTAL" else GREEN_DARK, width=16)
            v.grid(row=0, column=col*2+1)
            setattr(self, attr, v)

        # ── Terms ─────────────────────────────────────────────────────────────
        tf = ttk.LabelFrame(inner, text="  Terms & Conditions (Editable)  ", padding=10)
        tf.pack(fill="x", padx=15, pady=8)
        self.txt_terms = tk.Text(tf, height=6, font=("Arial",9),
                                  relief="solid", bd=1, bg="white")
        self.txt_terms.pack(fill="x")
        self.txt_terms.insert("1.0",
            "1. Prices are valid for a limited period.\n"
            "2. Payment to be made as per agreed terms.\n"
            "3. Warranty as per manufacturer terms.\n"
            "4. All disputes subject to local jurisdiction only.")

        # ── Status bar ────────────────────────────────────────────────────────
        self.status_var = tk.StringVar(value="Ready")
        tk.Label(inner, textvariable=self.status_var, font=("Arial",9),
                 bg="#e8f5e9", fg=GREEN_DARK, anchor="w",
                 relief="sunken", bd=1).pack(fill="x", padx=15, pady=2)

        # ── Action buttons ────────────────────────────────────────────────────
        bf = tk.Frame(inner, bg=BG); bf.pack(pady=12)
        for txt, cmd, clr in [
            ("  💾  Save Bill + PDF + Excel + Google Sheet  ", self._save_bill,  GREEN_DARK),
            ("  🖨  PDF Only  ",                               self._pdf_only,   ORANGE),
            ("  🗑  Clear Form  ",                             self._clear,      "#777777"),
        ]:
            tk.Button(bf, text=txt, font=("Arial",10,"bold"), bg=clr, fg="white",
                      relief="flat", padx=10, pady=9, cursor="hand2",
                      command=cmd).pack(side="left", padx=8)

    # ── Calculate ──────────────────────────────────────────────────────────────
    def _calc(self):
        try:
            qty=float(self.e_qty.get() or 0); rate=float(self.e_rate.get() or 0)
            cp =float(self.e_cgst.get() or 0); sp =float(self.e_sgst.get() or 0)
            tax=qty*rate; ca=tax*cp/100; sa=tax*sp/100; tot=tax+ca+sa
            self.lbl_tax.config(text=f"Rs. {tax:,.2f}")
            self.lbl_cgst.config(text=f"Rs. {ca:,.2f}")
            self.lbl_sgst.config(text=f"Rs. {sa:,.2f}")
            self.lbl_total.config(text=f"Rs. {tot:,.2f}")
        except ValueError:
            messagebox.showerror("Error","Enter valid numbers for Qty, Rate and GST %")

    def _get_data(self):
        try:
            qty=float(self.e_qty.get() or 0); rate=float(self.e_rate.get() or 0)
            cp =float(self.e_cgst.get() or 0); sp =float(self.e_sgst.get() or 0)
            tax=qty*rate; ca=tax*cp/100; sa=tax*sp/100; tt=ca+sa; tot=tax+tt
            return {
                "bill_no":        self.e_billno.get(),
                "date":           self.e_date.get(),
                "customer_name":  self.e_cname.get(),
                "address":        self.e_addr.get(),
                "phone":          self.e_phone.get(),
                "gstin":          self.e_cgstin.get(),
                "place_of_supply":self.e_pos.get(),
                "items_desc":     self.txt_desc.get("1.0","end").strip(),
                "hsn":            self.e_hsn.get(),
                "qty":qty,"rate":rate,"taxable_value":tax,
                "cgst_pct":cp,"cgst_amt":ca,
                "sgst_pct":sp,"sgst_amt":sa,
                "total_tax":tt,"total":tot,
                "terms":          self.txt_terms.get("1.0","end").strip(),
            }
        except ValueError:
            messagebox.showerror("Error","Fill Qty, Rate and GST % correctly.")
            return None

    def _save_bill(self):
        if not self.e_cname.get().strip():
            messagebox.showerror("Required","Customer Name is required!"); return
        d = self._get_data()
        if not d: return
        self._calc()

        # 1. Save to local Excel
        self.status_var.set("Saving to Excel...")
        self.update()
        row_data = save_bill_to_excel(d)

        # 2. Generate PDF
        self.status_var.set("Generating PDF...")
        self.update()
        pdf = f'Bill_{d["bill_no"]}_{d["customer_name"].replace(" ","_")}.pdf'
        generate_pdf(d, self.cfg, pdf)

        # 3. Google Sheets (background thread so UI doesn't freeze)
        def gs_sync():
            def cb(msg): self.status_var.set(msg)
            sync_to_gsheet(self.cfg, row_data, cb)
        if self.cfg.get("gsheet_enabled"):
            self.status_var.set("Syncing to Google Sheets...")
            self.update()
            threading.Thread(target=gs_sync, daemon=True).start()
        else:
            self.status_var.set(f"✅ Done!  Excel: {BILLS_FILE}  |  PDF: {pdf}")

        messagebox.showinfo("Bill Saved!",
            f"✅ Bill #{d['bill_no']} saved!\n\n"
            f"📊 Excel: {BILLS_FILE}\n"
            f"🖨  PDF: {pdf}\n"
            f"{'📡 Syncing to Google Sheets...' if self.cfg.get('gsheet_enabled') else '📡 Google Sheets: Not connected'}")

        # Increment bill number
        self.e_billno.delete(0,tk.END)
        self.e_billno.insert(0, str(int(d["bill_no"])+1))
        self._refresh_ledger()

    def _pdf_only(self):
        d = self._get_data()
        if not d: return
        self._calc()
        path = filedialog.asksaveasfilename(
            defaultextension=".pdf", filetypes=[("PDF","*.pdf")],
            initialfile=f'Bill_{d["bill_no"]}_{d["customer_name"].replace(" ","_")}.pdf')
        if path:
            generate_pdf(d, self.cfg, path)
            messagebox.showinfo("Done", f"PDF saved:\n{path}")

    def _clear(self):
        for e in [self.e_cname,self.e_phone,self.e_addr,self.e_cgstin]: e.delete(0,tk.END)
        self.e_pos.delete(0,tk.END); self.e_pos.insert(0,"Your State (Code)")
        self.e_date.delete(0,tk.END); self.e_date.insert(0,datetime.today().strftime("%d-%m-%Y"))
        self.e_qty.delete(0,tk.END); self.e_qty.insert(0,"1.00")
        self.e_rate.delete(0,tk.END)
        for lbl in [self.lbl_tax,self.lbl_cgst,self.lbl_sgst,self.lbl_total]:
            lbl.config(text="Rs. 0.00")
        self.status_var.set("Form cleared – ready for new bill")

    # ══════════════════════════════════════════════════════════════════════════
    #  COMPANY SETTINGS TAB
    # ══════════════════════════════════════════════════════════════════════════
    def _build_company_tab(self):
        f = self.tab_company
        lf = ttk.LabelFrame(f, text="  Company Details – All Editable  ", padding=15)
        lf.pack(fill="both", expand=True, padx=20, pady=15)
        lf.columnconfigure(1,weight=1); lf.columnconfigure(3,weight=1)

        fields = [
            ("Company / Trade Name","company_name", 0,0),
            ("Legal Name (Prop.)",  "legal_name",   0,2),
            ("Alt. Trade Name",     "trade_name",   1,0),
            ("GSTIN",               "gstin",        1,2),
            ("Address",             "address",      2,0),
            ("UDYAM No.",           "udyam",        2,2),
            ("Phone",               "phone",        3,0),
            ("Email",               "email",        3,2),
            ("Bank Name",           "bank_name",    4,0),
            ("Account Name",        "account_name", 4,2),
            ("Account No.",         "account_no",   5,0),
            ("Account Type",        "account_type", 5,2),
            ("IFSC Code",           "ifsc",         6,0),
            ("Branch",              "branch",       6,2),
            ("CGST % (default)",    "cgst_rate",    7,0),
            ("SGST % (default)",    "sgst_rate",    7,2),
        ]
        self.cfg_entries = {}
        for label,key,row,col in fields:
            tk.Label(lf, text=label+":", font=("Arial",9,"bold"),
                     fg=GREEN_DARK, bg=BG).grid(row=row, column=col, sticky="w", padx=8, pady=5)
            e = tk.Entry(lf, font=("Arial",10), relief="solid", bd=1, bg="white")
            e.insert(0, self.cfg.get(key,""))
            e.grid(row=row, column=col+1, sticky="ew", padx=8, pady=5)
            self.cfg_entries[key] = e

        # Logo row
        tk.Label(lf, text="Logo File:", font=("Arial",9,"bold"),
                 fg=GREEN_DARK, bg=BG).grid(row=8,column=0,sticky="w",padx=8,pady=5)
        self.logo_entry = tk.Entry(lf, font=("Arial",10), relief="solid", bd=1, bg="white")
        self.logo_entry.insert(0, self.cfg.get("logo_path",""))
        self.logo_entry.grid(row=8, column=1, sticky="ew", padx=8)
        tk.Button(lf, text="Browse…", font=("Arial",9), bg=GREEN_DARK, fg="white",
                  relief="flat", command=self._pick_logo).grid(row=8,column=2,padx=5)

        tk.Button(f, text="  💾  Save All Settings  ", font=("Arial",12,"bold"),
                  bg=GREEN_DARK, fg="white", relief="flat", padx=20, pady=9,
                  cursor="hand2", command=self._save_company).pack(pady=14)

    def _pick_logo(self):
        p = filedialog.askopenfilename(filetypes=[("Images","*.png *.jpg *.jpeg")])
        if p: self.logo_entry.delete(0,tk.END); self.logo_entry.insert(0,p)

    def _save_company(self):
        for k,e in self.cfg_entries.items(): self.cfg[k] = e.get()
        self.cfg["logo_path"] = self.logo_entry.get()
        save_config(self.cfg)
        self.title("GST Billing – "+self.cfg["company_name"])
        messagebox.showinfo("Saved","✅ Company settings saved!")

    # ══════════════════════════════════════════════════════════════════════════
    #  GOOGLE SHEETS TAB
    # ══════════════════════════════════════════════════════════════════════════
    def _build_gsheet_tab(self):
        f = self.tab_gsheet

        # Instructions panel
        info_frame = tk.Frame(f, bg="#fff8e1", bd=1, relief="solid")
        info_frame.pack(fill="x", padx=20, pady=12)
        instructions = (
            "HOW TO CONNECT TO GOOGLE SHEETS (One-Time Setup)\n\n"
            "Step 1:  Go to  https://console.cloud.google.com/\n"
            "Step 2:  Create a new project (e.g. MyBillingProject)\n"
            "Step 3:  Enable  'Google Sheets API'  and  'Google Drive API'\n"
            "Step 4:  Go to  IAM & Admin → Service Accounts → Create Service Account\n"
            "Step 5:  Create a JSON key → Download it\n"
            "Step 6:  Rename the downloaded file to  google_credentials.json\n"
            "         and place it in the SAME folder as billing_app.py\n\n"
            "Step 7:  Create a new Google Sheet and copy its ID from the URL:\n"
            "         https://docs.google.com/spreadsheets/d/  SHEET_ID_HERE  /edit\n\n"
            "Step 8:  Open the Google Sheet → Share → Paste the Service Account email\n"
            "         (shown in the JSON file as  'client_email')  → Give Editor access\n\n"
            "Step 9:  Paste the Sheet ID below and Enable sync → Save\n"
            "         Every new bill will now auto-update your Google Sheet instantly!"
        )
        tk.Label(info_frame, text=instructions, font=("Courier",9),
                 bg="#fff8e1", fg="#5d4037", justify="left",
                 anchor="w").pack(padx=15, pady=10, fill="x")

        # Settings panel
        sf = ttk.LabelFrame(f, text="  Google Sheets Connection Settings  ", padding=15)
        sf.pack(fill="x", padx=20, pady=10)
        sf.columnconfigure(1, weight=1)

        tk.Label(sf, text="Google Sheet ID:", font=("Arial",10,"bold"),
                 bg=BG, fg=GREEN_DARK).grid(row=0, column=0, sticky="w", padx=8, pady=6)
        self.gs_id_entry = tk.Entry(sf, font=("Arial",10), relief="solid", bd=1, bg="white")
        self.gs_id_entry.insert(0, self.cfg.get("gsheet_id",""))
        self.gs_id_entry.grid(row=0, column=1, sticky="ew", padx=8, pady=6)

        tk.Label(sf, text="Sheet Tab Name:", font=("Arial",10,"bold"),
                 bg=BG, fg=GREEN_DARK).grid(row=1, column=0, sticky="w", padx=8, pady=6)
        self.gs_tab_entry = tk.Entry(sf, font=("Arial",10), relief="solid", bd=1, bg="white")
        self.gs_tab_entry.insert(0, self.cfg.get("gsheet_tab","Bills"))
        self.gs_tab_entry.grid(row=1, column=1, sticky="ew", padx=8, pady=6)

        tk.Label(sf, text="Credentials File:", font=("Arial",10,"bold"),
                 bg=BG, fg=GREEN_DARK).grid(row=2, column=0, sticky="w", padx=8, pady=6)
        cred_status = "✅ google_credentials.json FOUND" if os.path.exists(CREDS_FILE) \
                      else "❌ google_credentials.json NOT FOUND – download from Google Cloud"
        cred_color  = GREEN_DARK if os.path.exists(CREDS_FILE) else "#c00000"
        self.gs_cred_lbl = tk.Label(sf, text=cred_status, font=("Arial",9),
                                     bg=BG, fg=cred_color)
        self.gs_cred_lbl.grid(row=2, column=1, sticky="w", padx=8)

        # Enable toggle
        self.gs_enabled = tk.BooleanVar(value=bool(self.cfg.get("gsheet_enabled", False)))
        tk.Checkbutton(sf, text="Enable Google Sheets Sync",
                       variable=self.gs_enabled, font=("Arial",10,"bold"),
                       bg=BG, fg=GREEN_DARK, activebackground=BG,
                       selectcolor="white").grid(row=3, column=0, columnspan=2,
                                                  sticky="w", padx=8, pady=8)

        btn_frame = tk.Frame(f, bg=BG); btn_frame.pack(pady=8)
        tk.Button(btn_frame, text="  💾  Save Google Sheets Settings  ",
                  font=("Arial",11,"bold"), bg=GREEN_DARK, fg="white",
                  relief="flat", padx=14, pady=8, cursor="hand2",
                  command=self._save_gsheet).pack(side="left", padx=8)
        tk.Button(btn_frame, text="  🔌  Test Connection  ",
                  font=("Arial",11,"bold"), bg=ORANGE, fg="white",
                  relief="flat", padx=14, pady=8, cursor="hand2",
                  command=self._test_gsheet).pack(side="left", padx=8)

        self.gs_status = tk.Label(f, text="", font=("Arial",10,"bold"),
                                   bg=BG, fg=GREEN_DARK)
        self.gs_status.pack(pady=6)

    def _save_gsheet(self):
        self.cfg["gsheet_id"]      = self.gs_id_entry.get().strip()
        self.cfg["gsheet_tab"]     = self.gs_tab_entry.get().strip() or "Bills"
        self.cfg["gsheet_enabled"] = self.gs_enabled.get()
        save_config(self.cfg)
        # Refresh cred status
        cred_ok = os.path.exists(CREDS_FILE)
        self.gs_cred_lbl.config(
            text="✅ google_credentials.json FOUND" if cred_ok
                 else "❌ google_credentials.json NOT FOUND",
            fg=GREEN_DARK if cred_ok else "#c00000")
        messagebox.showinfo("Saved",
            "Google Sheets settings saved!\n\n"
            + ("✅ Sync is ENABLED – bills will auto-update your sheet."
               if self.gs_enabled.get()
               else "ℹ Sync is DISABLED."))

    def _test_gsheet(self):
        self.gs_status.config(text="Testing connection...", fg=GREEN_DARK)
        self.update()
        def run_test():
            try:
                import gspread
                from google.oauth2.service_account import Credentials
                if not os.path.exists(CREDS_FILE):
                    raise FileNotFoundError("google_credentials.json not found")
                scopes = ["https://spreadsheets.google.com/feeds",
                          "https://www.googleapis.com/auth/drive"]
                creds = Credentials.from_service_account_file(CREDS_FILE, scopes=scopes)
                gc = gspread.authorize(creds)
                sh = gc.open_by_key(self.cfg["gsheet_id"])
                self.gs_status.config(
                    text=f"✅ Connected!  Sheet: '{sh.title}'  — Ready to sync!",
                    fg=GREEN_DARK)
            except Exception as e:
                self.gs_status.config(text=f"❌ Error: {e}", fg="#c00000")
        threading.Thread(target=run_test, daemon=True).start()

    # ══════════════════════════════════════════════════════════════════════════
    #  BILLS LEDGER TAB
    # ══════════════════════════════════════════════════════════════════════════
    def _build_ledger_tab(self):
        f = self.tab_ledger
        top = tk.Frame(f, bg=BG); top.pack(fill="x", padx=10, pady=8)
        tk.Label(top, text="All Bills Ledger", font=("Arial",13,"bold"),
                 bg=BG, fg=GREEN_DARK).pack(side="left")
        tk.Button(top, text="🔄 Refresh", font=("Arial",9), bg=GREEN_DARK, fg="white",
                  relief="flat", padx=8, command=self._refresh_ledger).pack(side="left",padx=8)
        tk.Button(top, text="📂 Open Excel", font=("Arial",9), bg=ORANGE, fg="white",
                  relief="flat", padx=8, command=self._open_excel).pack(side="left")
        tk.Label(top, text="  Search:", font=("Arial",9,"bold"),
                 bg=BG, fg=GREEN_DARK).pack(side="left", padx=(16,4))
        self.search_var = tk.StringVar()
        self.search_var.trace("w", lambda *a: self._filter_ledger())
        tk.Entry(top, textvariable=self.search_var, width=22, font=("Arial",10),
                 relief="solid", bd=1).pack(side="left")

        cols = ("Bill No","Date","Customer","Phone","Taxable","CGST","SGST","Total")
        self.tree = ttk.Treeview(f, columns=cols, show="headings", height=26)
        for col, w in zip(cols, [70,95,200,110,115,95,95,125]):
            self.tree.heading(col, text=col)
            self.tree.column(col, width=w, anchor="center")
        sy = ttk.Scrollbar(f, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=sy.set)
        sy.pack(side="right", fill="y")
        self.tree.pack(fill="both", expand=True, padx=10, pady=4)
        self.tree.tag_configure("odd",  background=ROW_ALT)
        self.tree.tag_configure("even", background="white")
        self.all_rows = []
        self._refresh_ledger()

    def _refresh_ledger(self):
        for r in self.tree.get_children(): self.tree.delete(r)
        self.all_rows = []
        if not os.path.exists(BILLS_FILE): return
        wb = openpyxl.load_workbook(BILLS_FILE, data_only=True)
        ws = wb["Bills"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]: self.all_rows.append(row)
        self._populate_ledger(self.all_rows)

    def _populate_ledger(self, rows):
        for r in self.tree.get_children(): self.tree.delete(r)
        for i, row in enumerate(rows):
            vals = (row[0], row[1], row[2], row[4] or "",
                    f"Rs.{float(row[11] or 0):,.2f}",
                    f"Rs.{float(row[13] or 0):,.2f}",
                    f"Rs.{float(row[15] or 0):,.2f}",
                    f"Rs.{float(row[16] or 0):,.2f}")
            self.tree.insert("","end", values=vals, tags=("odd" if i%2 else "even",))

    def _filter_ledger(self):
        q = self.search_var.get().lower()
        self._populate_ledger(
            [r for r in self.all_rows if any(q in str(v).lower() for v in r)]
            if q else self.all_rows)

    def _open_excel(self):
        if not os.path.exists(BILLS_FILE):
            messagebox.showinfo("Info","No bills saved yet."); return
        import subprocess, sys
        if sys.platform=="win32":    os.startfile(BILLS_FILE)
        elif sys.platform=="darwin": subprocess.call(["open", BILLS_FILE])
        else:                        subprocess.call(["xdg-open", BILLS_FILE])


if __name__ == "__main__":
    App().mainloop()
